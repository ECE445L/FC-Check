import time
import requests
import logging, sys, os 
import importlib.util, inspect
from typing import List, Tuple
import openpyxl

# TUI Library
from rich.prompt import Prompt
from rich import print
from rich.panel import Panel
from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt,IntPrompt

# DEBUG Flag that utilizes logging to print to stdout debug info
DEBUG = False 

APP_ID    = "846098"
CLIENT_ID = "Iv1.f82874f4aab1a1f8"

# The goal of this script is to use the GH API in order to check if the student has a repo
# With at least one commit

def read_excel_to_arrays(filepath):
  """
  Reads the first two columns of the first worksheet in an Excel file 
  and stores them in a pair of arrays.

  Args:
      filepath: The path to the Excel file.
  """
  # Open the workbook
  wb = openpyxl.load_workbook(filename=filepath)

  # Get the first sheet
  sheet = wb.active

  # Initialize empty arrays
  col1 = []
  col2 = []

  # Loop through rows (starting from 1 to skip headers)
  for row in sheet.iter_rows(min_row=2):
    # Append data to respective arrays
    col1.append(row[0].value)
    col2.append(row[1].value)

  return (col1, col2)

def write_data_to_excel(data_arrays, filename):
  """
  Writes four arrays of data as strings to the columns of a new Excel file.

  Args:
      data_arrays: A list containing four arrays of data to be written.
      filename: The desired name of the output Excel file.
  """
  # Create a new workbook
  wb = Workbook()
  
  # Get the worksheet
  ws = wb.active

  ws.append(["Name", "EID", "Has Repo", "Did Commit"])

  # Write data to rows
  for row_index, row_data in enumerate(zip(*data_arrays)):
    ws.append(list(row_data))

  # Save the workbook
  wb.save(filename)


"""
r = requests.get('https://github.com/timeline.json')
print( r.json() )
"""

"""
url = 'https://www.googleapis.com/qpxExpress/v1/trips/search?key=mykeyhere'
payload = open("request.json")
headers = {'content-type': 'application/json', 'Accept-Charset': 'UTF-8'}

r = requests.post(url, data=payload, headers=headers)

"""

def get_gh_auth() -> str:
  """
  Attempts to get and return a Github user token
  """

  #Request a token to identify ourself
  url = 'https://github.com/login/device/code' + f'?client_id={CLIENT_ID}'
  r = requests.post(url)

  fields = {elem.split("=")[0]: elem.split("=")[1] for elem in str(r.content)[2:-1].split("&")}

  token = fields["device_code"]
  user_code = fields["user_code"]

  print(f"")
  print(f"------------------------------------")
  print(f"You must login to github to continue")
  print(f"Enter the code '{user_code}'")
  print(f"at https://github.com/login/device")
  print(f"Note: UC = '{user_code}'")
  print(f"------------------------------------")

  script_dir = os.path.dirname(os.path.abspath(__file__))
  gh_token_path = os.path.join(script_dir, "gh_token")
  
  gh_token = ""
  is_auth = False
  # Awaits for user to autheticate the app on GH. Prints a spinner animation
  console = Console()
  with console.status("[bold green]Waiting for authorization...") as status:


    # If the app token exists on disk, read it
    if os.path.exists(gh_token_path):
      with open(gh_token_path, "r") as file:
        gh_token = file.read()
        url = 'https://api.github.com/user' 
        headers = {"Accept":"application/vnd.github+json", "Authorization": f"Bearer {gh_token}", "X-GitHub-Api-Version":"2022-11-28"}
        r = requests.get(url, headers=headers)
        if r.status_code == 200: 
          is_auth = True
        else:
          print("\nToken on disk mismatch. Please visit link above and authorize\n")
    
    # Otherwise if the token doesnt exist or was invalid, ask for a new one
    while(not is_auth):

      #Check that the token has been acessed
      url = 'https://github.com/login/oauth/access_token' 
      parameters = {"client_id": CLIENT_ID, "device_code": token, "grant_type": "urn:ietf:params:oauth:grant-type:device_code"}
      r = requests.post(url, parameters)
      
      if r.status_code == 200:
        if "authorization_pending" not in str(r.content):
          fields = {elem.split("=")[0]: elem.split("=")[1] for elem in str(r.content)[2:-1].split("&")}
          logging.debug(fields)
          logging.debug(fields["access_token"])
          gh_token = fields["access_token"]
          is_auth = True
          break
      else:
        print(f"Invalid status seen? See content/status_code: {r.content}, {r.status_code}\n")
      time.sleep(5)

  # Writes the token to a file to be reused
  with open(gh_token_path, "w") as file:
        # Storing a token in plaintext isnt the smartest idea, but uh, who cares.
        file.write(gh_token)

  return gh_token

def get_gh_classroom(token:str):
  """
  Shows the user all github classrooms they are a part of to have them select one 
  """

  url = f'https://api.github.com/classrooms' 
  parameters = {"page":1, "per_page":100}
  headers = {"Accept":"application/vnd.github+json", "Authorization": f"Bearer {token}", "X-GitHub-Api-Version":"2022-11-28"}
  r = requests.get(url, parameters, headers=headers)

  if(r.status_code != 200):
    print("Error:", r.content, r.status_code)
    exit()

  rjson = r.json()

  # Prints a listing of all the classrooms the user is part of in a table
  console = Console()
  table = Table(show_header=True, header_style="bold magenta")
  table.add_column("Selection", style="dim")
  table.add_column("Classroom Name")
  table.add_column("Classroom ID",justify="right")

  for i, rdict in enumerate(rjson):
    table.add_row(str(i),rdict['name'],str(rdict['id']))
  console.print(table)

  # User selects a classroom
  idx = IntPrompt.ask("Select a classroom", choices=[str(i) for i in range(len(rjson))])

  return rjson[idx]['id']

def get_gh_assignment(token:str, classroom_id : str):
  """
  Shows the user all github assignments a classrooms has to select one 
  """

  url = f'https://api.github.com/classrooms/{classroom_id}/assignments' 
  parameters = {"page":1, "per_page":100}
  headers = {"Accept":"application/vnd.github+json", "Authorization": f"Bearer {token}", "X-GitHub-Api-Version":"2022-11-28"}
  r = requests.get(url, parameters, headers=headers)

  if(r.status_code != 200):
    print("Error:", r.content, r.status_code)
    exit()

  rjson = r.json()

  # Prints a listing of all the assignments of the class in a table
  console = Console()
  table = Table(show_header=True, header_style="bold magenta")
  table.add_column("Selection", style="dim")
  table.add_column("Assignment Title")
  table.add_column("Assignment ID",justify="right")

  for i, rdict in enumerate(rjson):
    table.add_row(str(i),rdict['title'],str(rdict['id']))
  console.print(table)

  idx = IntPrompt.ask("Select a assignment", choices=[str(i) for i in range(len(rjson))])

  return rjson[idx]['id']

def get_gh_assignment_info(token: str, assignment_id: str):
  """
  Attempts to load an assignment by id
  """

  url = f'https://api.github.com/assignments/{assignment_id}' 
  parameters = {"page":1, "per_page":100}
  headers = {"Accept":"application/vnd.github+json", "Authorization": f"Bearer {token}", "X-GitHub-Api-Version":"2022-11-28"}
  r = requests.get(url, parameters, headers=headers)

  rjson = r.json()
  logging.debug(r.content, r.status_code, rjson)

  print(Panel(
    f"[bold italic red]Assignment[/]: '{rjson['title']}'\n"
    f"[bold]Invite Link[/]: '{rjson['invite_link']}'\n"
    f"[bold]Deadline[/]: '{rjson['deadline']}'\n"
    f"[bold magenta]Repo count[/]: '{rjson['accepted']}'\n"
    f"[bold]==========\n"
    f"[bold green]Classroom[/]: '{rjson['classroom']['url']}'\n"
    f"[bold green]Starter Code[/]: '{rjson['starter_code_repository']['html_url']}'\n"
    , title="Assignment"))

  url = f'https://api.github.com/assignments/{assignment_id}/accepted_assignments' 
  parameters = {"page":1, "per_page":100}
  headers = {"Accept":"application/vnd.github+json", "Authorization": f"Bearer {token}", "X-GitHub-Api-Version":"2022-11-28"}
  r = requests.get(url, parameters, headers=headers)
  rjson = r.json()
  # logging.log(r.content, r.status_code, rjson)


def load_tests():
    test_modules = []
    script_dir = os.path.dirname(os.path.abspath(__file__))
    tests_dir = os.path.join(script_dir, "tests")
    # Iterate through files in the tests directory
    for root, dirs, files in os.walk(tests_dir):
        for file in files:
            if file.endswith(".py") and file != "__init__.py":  # Exclude __init__.py
                file_path = os.path.join(root, file)
                module_name = os.path.splitext(file)[0]  # Module name is file name without extension

                # Dynamically import the module
                spec = importlib.util.spec_from_file_location(module_name, file_path)
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)

                # Check if the module contains a Test class
                if hasattr(module, "Test") and inspect.isclass(module.Test):
                    test_class = module.Test
                    # Instantiate the Test class to get metadata
                    test_instance = test_class()
                    # Append the test module along with its metadata if enabled
                    if(test_instance.enabled):
                      test_modules.append((module, test_instance))

    return test_modules

def list_enabled_tests(loaded_tests):
  console = Console()
  console.print("\n[bold green italic]The following tests will be run on each students repo")
  table = Table(show_header=True, header_style="bold magenta")
  table.add_column("Test Name", style="dim")
  table.add_column("Description")

  # List the test names and description
  for test_module, test_instance in test_modules:
    table.add_row(test_instance.name,test_instance.description)
  console.print(table)

if __name__ == "__main__":
	
  if(DEBUG): logging.basicConfig(stream=sys.stderr, level=logging.DEBUG)
  

  print(Panel.fit("ECE-445L Fast Code Check"))

  # Authetnicates GH user and fetcehs classroom and assignment

  ghu_token = get_gh_auth()
  
  gh_c_id = get_gh_classroom(ghu_token)

  gh_a_id = get_gh_assignment(ghu_token, gh_c_id)

  get_gh_assignment_info(ghu_token, gh_a_id)

  # Load all the test modules along with their metadata
  test_modules = load_tests()
  list_enabled_tests(test_modules)

  in_start_test = Prompt.ask("[bold] Run tests?", choices=['y','n'])

  if(in_start_test == 'y'):
    console = Console()
    print("\n\n\n")
    print(Panel.fit("[bold]Starting tests"))
    with console.status("[bold green] Running Tests") as all_tests_status:
      for test_module, test_instance in test_modules:
        print(f"\nRunning test: '{test_instance.name}'")

        # TODO: Pass in relevant info to the test. Get a CSV output that can be recorded.
        test_result = test_instance.run_test()
        
        #TODO: Write result to a spreadsheet